import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.cell.cell import MergedCell
import pandas as pd
import os
import re
import math
import sys
import warnings

# Отключаем лишние предупреждения в консоли
warnings.simplefilter("ignore")

# НАСТРОЙКИ ОКНА
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.geometry("750x800") 
app.title("Auto-invoice v5.0 - The Smart Aggregator")

# ПЕРЕМЕННЫЕ ДЛЯ ПУТЕЙ
bi_files = []      
template_files = [] 
token_files = []

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
def update_status_label(label, file_list):
    if not file_list:
        label.configure(text="Файлы не выбраны", text_color="gray")
    else:
        names = [os.path.basename(f) for f in file_list]
        text = ", ".join(names)
        if len(text) > 75:
            text = text[:72] + "..."
        label.configure(text=f"Выбрано ({len(file_list)}): {text}", text_color="#a3c1e0")

def clear_bi():
    global bi_files
    bi_files = []
    update_status_label(lbl_bi, bi_files)
    log_text.insert("end", "[ОТМЕНА] Выбор BI-отчетов сброшен.\n\n")

def clear_tokens():
    global token_files
    token_files = []
    update_status_label(lbl_tokens, token_files)
    log_text.insert("end", "[ОТМЕНА] Выбор Токенов сброшен.\n\n")

def clear_templates():
    global template_files
    template_files = []
    update_status_label(lbl_tpl, template_files)
    log_text.insert("end", "[ОТМЕНА] Выбор Шаблонов сброшен.\n\n")

def load_bi_reports():
    global bi_files
    new_files = filedialog.askopenfilenames(
        title="Выберите все 3 BI-отчета (ISS, ACQ, AUTH)",
        filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv")]
    )
    if new_files:
        bi_files = list(new_files)
        update_status_label(lbl_bi, bi_files)
        log_text.insert("end", f"[OK] Выбрано BI-отчетов: {len(bi_files)}\n")

def load_token_reports():
    global token_files
    new_files = filedialog.askopenfilenames(
        title="Выберите отчеты по токенам (.xlsx)",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if new_files:
        token_files = list(new_files)
        update_status_label(lbl_tokens, token_files)
        log_text.insert("end", f"[OK] Выбрано отчетов по токенам: {len(token_files)}\n")

def load_templates():
    global template_files
    new_files = filedialog.askopenfilenames(
        title="Выберите ВСЕ Шаблоны счетов",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if new_files:
        template_files = list(new_files)
        update_status_label(lbl_tpl, template_files)
        log_text.insert("end", f"[OK] Выбрано шаблонов: {len(template_files)}\n")

def calculate_billed_tokens(raw_tokens):
    try:
        val = int(float(raw_tokens))
        if val <= 0: return 0
        return math.ceil(val / 10)
    except: return 0

def heal_excel_file_if_possible(filepath, log_widget):
    if not sys.platform.startswith('win'): return False
    try:
        import win32com.client as win32
        log_widget.insert("end", f"  [*] (Windows) Автоматически лечу структуру файла {os.path.basename(filepath)}...\n")
        app.update()
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(filepath))
        wb.Save()
        wb.Close()
        excel.Quit()
        return True
    except Exception as e:
        try: excel.Quit() 
        except: pass
        return False

# СВЕРХТОЧНАЯ ОЧИСТКА ИМЕН
def clean_bank_name(name):
    if pd.isna(name) or not name: return ""
    name = str(name).lower()
    name = re.sub(r'\s*\((mc|visa|mastercard)\)', '', name)
    # Удаляем формы собственности (CY и UK больше не удаляем, чтобы не слить Decta)
    name = re.sub(r'\b(ltd|limited|uab|a/s|as|sia|llc|inc)\b', '', name)
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

def extract_template_client_name(filename):
    name = re.sub(r'^\d{4}-\d{2}_', '', filename)
    name = re.split(r'_Invoice', name, flags=re.IGNORECASE)[0]
    return name

def match_bank_to_template(report_bank_name, tpl_filename):
    b_low = str(report_bank_name).lower()
    f_low = str(tpl_filename).lower()
    
    clean_b = clean_bank_name(report_bank_name)
    clean_t = clean_bank_name(extract_template_client_name(tpl_filename))
    
    if clean_b and (clean_b in clean_t or clean_t in clean_b):
        # Жесткая защита от слияния Decta CY, UK, LU с обычным Decta
        if "decta" in clean_b and "decta" in clean_t:
            for suffix in ['cy', 'uk', 'lu', 'ie']:
                if (suffix in clean_b) != (suffix in clean_t):
                    return False

        is_mc = "(mc)" in b_low or "(mastercard)" in b_low
        is_visa = "(visa)" in b_low
        
        if is_mc: return "mc" in f_low or "mastercard" in f_low
        if is_visa: return "visa" in f_low
        return True
    return False

# --- УМНАЯ ЗАПИСЬ (ОБХОД ОШИБКИ MERGED CELLS) ---
def safe_write(sheet, row_idx, col_idx, value):
    cell = sheet.cell(row=row_idx, column=col_idx)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = value
                return
    else:
        cell.value = value

# БЕЗОПАСНОЕ ЧТЕНИЕ ЛЮБЫХ ФАЙЛОВ BI
def read_bi_file_robust(filepath):
    encodings = ['utf-8', 'cp1251', 'latin1']
    separators = [',', ';']
    for enc in encodings:
        for sep in separators:
            try:
                with open(filepath, 'r', encoding=enc, errors='ignore') as f:
                    lines = f.readlines()[:30]
                h_idx = -1
                for i, line in enumerate(lines):
                    if "Issuer" in line or "Acquirer" in line:
                        h_idx = i
                        break
                if h_idx != -1:
                    df = pd.read_csv(filepath, header=h_idx, encoding=enc, sep=sep, on_bad_lines='skip')
                    if not df.empty and len(df.columns) > 1: return df
            except: pass
    try:
        df_temp = pd.read_excel(filepath, nrows=30, header=None)
        h_idx = 0
        for i, row in df_temp.iterrows():
            rstr = " ".join([str(x) for x in row.values if pd.notna(x)]).lower()
            if "issuer" in rstr or "acquirer" in rstr:
                h_idx = i
                break
        df = pd.read_excel(filepath, header=h_idx)
        if not df.empty: return df
    except: pass
    return None

def generate_invoices():
    if not bi_files or not template_files:
        log_text.insert("end", "[ОШИБКА] Сначала загрузите BI-отчеты и Шаблоны!\n\n")
        return
    save_folder = filedialog.askdirectory(title="Выберите папку для сохранения счетов")
    if not save_folder: return
    
    log_text.insert("end", "\n====================================\n")
    log_text.insert("end", "[*] НАЧИНАЕМ СБОР ДАННЫХ...\n")
    app.update()

    try:
        iss_data = {}
        acq_data = {}
        auth_data = {}
        total_tokens = 0

        # ЧИТАЕМ ТОКЕНЫ
        for path in token_files:
            try:
                filename = os.path.basename(path).upper()
                month_match = re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s*\_?\s*(\d{4})', filename)
                target_date_str = f"{month_match.group(1)}{month_match.group(2)}" if month_match else None
                
                wb = openpyxl.load_workbook(path, data_only=True)
                for row in wb.active.iter_rows(min_row=2, values_only=True):
                    if len(row) > 11:
                        col_g = str(row[6]).strip().upper().replace(" ", "") if row[6] else ""
                        if target_date_str and target_date_str not in col_g: continue 
                        col_h, col_k, col_l = row[7], str(row[10]).strip().upper() if row[10] else "", str(row[11]).strip().upper() if row[11] else ""
                        if col_l == "NEW" or (col_l == "0" and col_k == "CHANGES!"):
                            total_tokens += calculate_billed_tokens(col_h)
                wb.close() 
            except Exception as e:
                log_text.insert("end", f"[!] Ошибка при чтении токенов: {e}\n")

        # ЧИТАЕМ BI-ОТЧЕТЫ
        for path in bi_files:
            filename = os.path.basename(path).upper()
            if "MOZART" in filename: continue
            df = read_bi_file_robust(path)
            if df is None or df.empty: continue
            df.columns = [str(c).strip().replace('\n', ' ') for c in df.columns]

            if "ISS" in filename and "AUTH" not in filename:
                log_text.insert("end", f"[*] Динамический парсинг ISS: {os.path.basename(path)}\n")
                
                # Ищем колонку Транзакций
                iss_col_txn, txn_col = None, None
                for i, col in enumerate(df.columns):
                    if 'transactions count' in str(col).lower():
                        txn_col = col
                        for j in range(i-1, -1, -1):
                            if 'issuer' in str(df.columns[j]).lower():
                                iss_col_txn = df.columns[j]
                                break
                        break
                
                # Ищем колонки Карт
                iss_col_cards, cards_per_col, cards_tot_col = None, None, None
                for i, col in enumerate(df.columns):
                    if 'card count in period' in str(col).lower():
                        cards_per_col = col
                        for j in range(i-1, -1, -1):
                            if 'issuer' in str(df.columns[j]).lower():
                                iss_col_cards = df.columns[j]
                                break
                    if 'count_cards' in str(col).lower():
                        cards_tot_col = col

                # Извлекаем данные
                if iss_col_txn and txn_col:
                    for _, r in df.dropna(subset=[iss_col_txn, txn_col]).iterrows():
                        b = str(r[iss_col_txn]).strip()
                        if b and b.lower() not in ['nan', 'issuer']:
                            if b not in iss_data: iss_data[b] = {"txn": 0, "cards_total": 0, "cards_period": 0}
                            try: iss_data[b]["txn"] = float(r[txn_col])
                            except: pass

                if iss_col_cards and cards_per_col:
                    for _, r in df.dropna(subset=[iss_col_cards, cards_per_col]).iterrows():
                        b = str(r[iss_col_cards]).strip()
                        if b and b.lower() not in ['nan', 'issuer']:
                            if b not in iss_data: iss_data[b] = {"txn": 0, "cards_total": 0, "cards_period": 0}
                            try: iss_data[b]["cards_period"] = float(r[cards_per_col])
                            except: pass
                            if cards_tot_col and pd.notna(r[cards_tot_col]):
                                try: iss_data[b]["cards_total"] = float(r[cards_tot_col])
                                except: pass

            elif "ACQ" in filename and "AUTH" not in filename:
                log_text.insert("end", f"[*] Динамический парсинг ACQ: {os.path.basename(path)}\n")
                
                acq_col_txn, txn_col = None, None
                for i, col in enumerate(df.columns):
                    if 'transaction count' in str(col).lower():
                        txn_col = col
                        for j in range(i-1, -1, -1):
                            if 'acquirer' in str(df.columns[j]).lower():
                                acq_col_txn = df.columns[j]
                                break
                        break
                        
                acq_col_chb, chb_col = None, None
                for i, col in enumerate(df.columns):
                    if 'chargeback count' in str(col).lower():
                        chb_col = col
                        for j in range(i-1, -1, -1):
                            if 'acquirer' in str(df.columns[j]).lower():
                                acq_col_chb = df.columns[j]
                                break
                        break

                if acq_col_txn and txn_col:
                    for _, r in df.dropna(subset=[acq_col_txn, txn_col]).iterrows():
                        b = str(r[acq_col_txn]).strip()
                        if b and b.lower() not in ['nan', 'acquirer name']:
                            if b not in acq_data: acq_data[b] = {"txn": 0, "chargebacks": 0}
                            try: acq_data[b]["txn"] = float(r[txn_col])
                            except: pass

                if acq_col_chb and chb_col:
                    for _, r in df.dropna(subset=[acq_col_chb, chb_col]).iterrows():
                        b = str(r[acq_col_chb]).strip()
                        if b and b.lower() not in ['nan', 'acquirer name']:
                            if b not in acq_data: acq_data[b] = {"txn": 0, "chargebacks": 0}
                            try: acq_data[b]["chargebacks"] = float(r[chb_col])
                            except: pass

            elif "AUTH" in filename:
                log_text.insert("end", f"[*] Динамический парсинг AUTH: {os.path.basename(path)}\n")
                for i, col in enumerate(df.columns):
                    if 'auth count' in str(col).lower():
                        is_acq = False
                        target_bank_col = None
                        for j in range(i-1, -1, -1):
                            prev_col = str(df.columns[j]).lower()
                            if 'acquirer' in prev_col:
                                is_acq = True
                                target_bank_col = df.columns[j]
                                break
                            elif 'issuer' in prev_col:
                                is_acq = False
                                target_bank_col = df.columns[j]
                                break
                        
                        if target_bank_col:
                            for _, r in df.dropna(subset=[target_bank_col, col]).iterrows():
                                b = str(r[target_bank_col]).strip()
                                if b and b.lower() not in ['nan', 'acquirer name', 'issuer']:
                                    if b not in auth_data: auth_data[b] = {"auth_acq": 0, "auth_iss": 0}
                                    if is_acq:
                                        try: auth_data[b]["auth_acq"] = float(r[col])
                                        except: pass
                                    else:
                                        try: auth_data[b]["auth_iss"] = float(r[col])
                                        except: pass

        all_banks = list(set(iss_data.keys()).union(set(acq_data.keys()), set(auth_data.keys())))
        app.update()

        found_iss_templates = set()
        found_acq_templates = set()

        log_text.insert("end", "\n[*] НАЧИНАЕМ ЗАПОЛНЕНИЕ ШАБЛОНОВ:\n")
        
        for tpl_path in template_files:
            tpl_filename = os.path.basename(tpl_path)
            possible_banks = [b for b in all_banks if match_bank_to_template(b, tpl_filename)]
            
            if not possible_banks:
                log_text.insert("end", f"  [ПРОПУСК] В '{tpl_filename}' клиент не найден!\n")
                continue
                
            bank_name = possible_banks[0] # Основное имя для отчета, но собирать будем ВСЕ совпадения!
            
            # --- СИНТЕЗАТОР ДАННЫХ (Объединяет TESLAPAY UAB и Teslapay) ---
            b_iss = {"txn": 0, "cards_total": 0, "cards_period": 0}
            b_acq = {"txn": 0, "chargebacks": 0}
            b_auth = {"auth_acq": 0, "auth_iss": 0}

            for pb in possible_banks:
                if pb in iss_data:
                    b_iss["txn"] += iss_data[pb].get("txn", 0)
                    b_iss["cards_total"] += iss_data[pb].get("cards_total", 0)
                    b_iss["cards_period"] += iss_data[pb].get("cards_period", 0)
                if pb in acq_data:
                    b_acq["txn"] += acq_data[pb].get("txn", 0)
                    b_acq["chargebacks"] += acq_data[pb].get("chargebacks", 0)
                if pb in auth_data:
                    b_auth["auth_acq"] += auth_data[pb].get("auth_acq", 0)
                    b_auth["auth_iss"] += auth_data[pb].get("auth_iss", 0)

            try:
                try:
                    wb_tpl = openpyxl.load_workbook(tpl_path)
                except Exception:
                    is_healed = heal_excel_file_if_possible(tpl_path, log_text)
                    if is_healed:
                        try: wb_tpl = openpyxl.load_workbook(tpl_path)
                        except: continue
                    else: continue
                
                target_sheet_name = None
                for name in wb_tpl.sheetnames:
                    if "annex" in name.lower() or "pielikums" in name.lower():
                        target_sheet_name = name
                        break
                if not target_sheet_name: target_sheet_name = wb_tpl.sheetnames[-1]
                    
                sheet_tpl = wb_tpl[target_sheet_name]
                
                is_acq = False
                c8_val = sheet_tpl["C8"].value if sheet_tpl["C8"].value else ""
                if "ACQUIRING" in str(c8_val).upper() or "ACQ" in tpl_filename.upper():
                    is_acq = True
                
                # --- ВСТАВКА (С защитой от строк тарифов) ---
                if is_acq:
                    found_acq_templates.add(bank_name)
                    log_text.insert("end", f"  [+] ACQ -> {bank_name} (Собраны данные из: {', '.join(possible_banks)})\n")

                    for r in range(8, 50):
                        b_val = str(sheet_tpl[f"B{r}"].value or "").strip()
                        c_val = str(sheet_tpl[f"C{r}"].value or "").strip().upper()
                        c_txt = c_val + " " + b_val.upper()
                        
                        is_data_row = bool(re.search(r'[a-zA-Z]', b_val)) or "ACTUAL" in c_val or "TOKEN" in c_val
                        if not is_data_row: continue

                        if "ACTUAL NUMBER OF AUTHORIZATIONS" in c_txt or "FM PROCESSED" in c_txt:
                            safe_write(sheet_tpl, r, 9, b_auth["auth_acq"])
                            log_text.insert("end", f"      -> Auth ({b_auth['auth_acq']}) в строку {r}\n")
                        elif "ACTUAL NUMBER OF TRANSACTIONS" in c_txt:
                            safe_write(sheet_tpl, r, 9, b_acq["txn"])
                            log_text.insert("end", f"      -> Trx ({b_acq['txn']}) в строку {r}\n")
                        elif ("CHARGEBACK" in c_txt or "CHARGEBACKS" in c_txt) and "GOOD FAITH" not in c_txt:
                            safe_write(sheet_tpl, r, 9, b_acq["chargebacks"])
                            log_text.insert("end", f"      -> Chargebacks ({b_acq['chargebacks']}) в строку {r}\n")

                else:
                    found_iss_templates.add(bank_name)
                    log_text.insert("end", f"  [+] ISS -> {bank_name} (Собраны данные из: {', '.join(possible_banks)})\n")

                    for r in range(8, 50):
                        b_val = str(sheet_tpl[f"B{r}"].value or "").strip()
                        c_val = str(sheet_tpl[f"C{r}"].value or "").strip().upper()
                        h_val = str(sheet_tpl[f"H{r}"].value or "").strip().upper()
                        c_txt = c_val + " " + b_val.upper()
                        
                        is_data_row = bool(re.search(r'[a-zA-Z]', b_val)) or "ACTUAL" in c_val or "TOKEN" in c_val or "TOKEN" in h_val
                        if not is_data_row: continue

                        if "ACTUAL NUMBER OF AUTHORIZATIONS" in c_txt or "FM PROCESSED" in c_txt:
                            safe_write(sheet_tpl, r, 9, b_auth["auth_iss"])
                            log_text.insert("end", f"      -> Auth ({b_auth['auth_iss']}) в строку {r}\n")
                        elif "ACTUAL NUMBER OF TRANSACTIONS" in c_txt or ("ISSUING TRANSACTIONS" in c_txt and "ACTUAL" in c_txt):
                            safe_write(sheet_tpl, r, 9, b_iss["txn"])
                            log_text.insert("end", f"      -> Trx ({b_iss['txn']}) в строку {r}\n")
                        elif ("CURRENT MONTH" in c_txt or "REPORT PERIOD" in c_txt) and "CARDS" in c_txt:
                            safe_write(sheet_tpl, r, 9, b_iss["cards_period"])
                            log_text.insert("end", f"      -> Cards Period ({b_iss['cards_period']}) в строку {r}\n")
                        elif ("WHOLE PERIOD" in c_txt or "TOTAL" in c_txt) and "CARDS" in c_txt:
                            safe_write(sheet_tpl, r, 9, b_iss["cards_total"])
                            log_text.insert("end", f"      -> Cards Total ({b_iss['cards_total']}) в строку {r}\n")
                        
                        if "TOKEN" in c_txt or "TOKEN" in h_val:
                            if total_tokens > 0:
                                safe_write(sheet_tpl, r, 9, total_tokens)
                                log_text.insert("end", f"      -> [✓] Токены ({total_tokens}) записаны в строку {r}\n")

                new_filename = f"READY_{tpl_filename}"
                wb_tpl.save(os.path.join(save_folder, new_filename))
                wb_tpl.close() 
                app.update()
                
            except Exception as e:
                log_text.insert("end", f"  [❌ ОШИБКА ШАБЛОНА] {tpl_filename} пропущен: {e}\n")
                continue

        # ОТЧЕТ
        missing_warnings = []
        for bank in iss_data.keys():
            if bank not in found_iss_templates and bank.upper() not in ["DECTA", "DECTA UK", "DECTA CY", "DECTA LU"]:
                missing_warnings.append(f"[{bank}] — забыт шаблон ISS")
        for bank in acq_data.keys():
            if bank not in found_acq_templates and bank.upper() not in ["DECTA LIMITED", "DECTA LIMITED CY"]:
                missing_warnings.append(f"[{bank}] — забыт шаблон ACQ")
                
        if missing_warnings:
            log_text.insert("end", "\n⚠️ [ВНИМАНИЕ] Пропущенные банки (в отчетах есть, а шаблонов нет):\n")
            for warning in missing_warnings:
                log_text.insert("end", f"  ❌ {warning}\n")

        log_text.insert("end", "\n[УСПЕХ] Обработка завершена!\n\n")
        log_text.see("end")

    except Exception as e:
        log_text.insert("end", f"\n[Critical Error] {e}\n\n")
        log_text.see("end")

# =======================
# ВИЗУАЛЬНАЯ ЧАСТЬ (UI)
# =======================
label_title = ctk.CTkLabel(app, text="Mass Invoice Generator v5.0", font=("Arial", 20, "bold"))
label_title.pack(pady=(15, 15))

frame_bi = ctk.CTkFrame(app, fg_color="transparent")
frame_bi.pack(pady=2)
btn_bi = ctk.CTkButton(frame_bi, text="1. Select 3 BI-reports (AUTH,ISS,ACQ)", command=load_bi_reports, width=260)
btn_bi.grid(row=0, column=0, padx=5)
btn_clear_bi = ctk.CTkButton(frame_bi, text="✖", width=35, command=clear_bi, fg_color="#c9302c", hover_color="#ac2925")
btn_clear_bi.grid(row=0, column=1, padx=5)

lbl_bi = ctk.CTkLabel(app, text="Файлы не выбраны", text_color="gray", font=("Arial", 11))
lbl_bi.pack(pady=(0, 10))

frame_tokens = ctk.CTkFrame(app, fg_color="transparent")
frame_tokens.pack(pady=2)
btn_tokens = ctk.CTkButton(frame_tokens, text="2. Select Token reports (.xlsx)", command=load_token_reports, width=260)
btn_tokens.grid(row=0, column=0, padx=5)
btn_clear_tokens = ctk.CTkButton(frame_tokens, text="✖", width=35, command=clear_tokens, fg_color="#c9302c", hover_color="#ac2925")
btn_clear_tokens.grid(row=0, column=1, padx=5)

lbl_tokens = ctk.CTkLabel(app, text="Файлы не выбраны", text_color="gray", font=("Arial", 11))
lbl_tokens.pack(pady=(0, 10))

frame_tpl = ctk.CTkFrame(app, fg_color="transparent")
frame_tpl.pack(pady=2)
btn_tpl = ctk.CTkButton(frame_tpl, text="3. Select ALL Templates", command=load_templates, width=260)
btn_tpl.grid(row=0, column=0, padx=5)
btn_clear_tpl = ctk.CTkButton(frame_tpl, text="✖", width=35, command=clear_templates, fg_color="#c9302c", hover_color="#ac2925")
btn_clear_tpl.grid(row=0, column=1, padx=5)

lbl_tpl = ctk.CTkLabel(app, text="Файлы не выбраны", text_color="gray", font=("Arial", 11))
lbl_tpl.pack(pady=(0, 15))

btn_gen = ctk.CTkButton(app, text="4. GENERATE INVOICES", command=generate_invoices, fg_color="#28a745", hover_color="#218838", width=305)
btn_gen.pack(pady=(5, 15))

log_text = ctk.CTkTextbox(app)
log_text.pack(padx=20, pady=(0, 20), fill="both", expand=True)
log_text.insert("end", "Программа готова! Загрузите нужные файлы.\n\n")

context_menu = tk.Menu(log_text, tearoff=False, font=("Arial", 11))
context_menu.add_command(label="Копировать", command=lambda: log_text.event_generate("<<Copy>>"))
def show_context_menu(event):
    try:
        log_text.get("sel.first", "sel.last")
        context_menu.tk_popup(event.x_root, event.y_root)
    except tk.TclError: pass 
log_text.bind("<Button-3>", show_context_menu)
if sys.platform == "darwin": log_text.bind("<Button-2>", show_context_menu)

app.mainloop()